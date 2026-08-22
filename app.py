import io
import math
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import pandas as pd
import streamlit as st
import folium
from streamlit_folium import st_folium
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

st.set_page_config(
    layout="wide",
    page_title="Petrol Station Trade Area Mapper",
    page_icon="⛽",
)

st.title("⛽ Petrol Station Trade Area Mapper (10 km Radius)")

# --- 1. SIDEBAR: DATA ENTRY ---
st.sidebar.header("📍 Site Entry Form")

entry_type = st.sidebar.selectbox(
    "Entry Type",
    ["Proposed Station", "Competitor Station", "Landmark / POI"],
)
name = st.sidebar.text_input("Station / Landmark Name", "Site Alpha")
lat = st.sidebar.number_input("Latitude", value=24.8607, format="%.5f")
lon = st.sidebar.number_input("Longitude", value=67.0011, format="%.5f")

# Initialize Session Data
if "elements" not in st.session_state:
    st.session_state.elements = [
        {
            "type": "Proposed Station",
            "name": "Proposed Subject Site",
            "omc": "SHELL",
            "lat": 24.8607,
            "lon": 67.0011,
            "pmg": 180,
            "hsd": 240,
            "hobc": 45,
        },
        {
            "type": "Competitor Station",
            "name": "North Metro Station",
            "omc": "TOTAL",
            "lat": 24.8950,
            "lon": 67.0250,
            "pmg": 120,
            "hsd": 150,
            "hobc": 15,
        },
        {
            "type": "Competitor Station",
            "name": "East Highway Hub",
            "omc": "PSO",
            "lat": 24.8300,
            "lon": 67.0500,
            "pmg": 210,
            "hsd": 300,
            "hobc": 30,
        },
        {
            "type": "Landmark / POI",
            "name": "KFC Drive-Thru & Hospital",
            "omc": "N/A",
            "lat": 24.8750,
            "lon": 67.0120,
            "pmg": 0,
            "hsd": 0,
            "hobc": 0,
        },
    ]

if entry_type in ["Proposed Station", "Competitor Station"]:
    omc = st.sidebar.text_input("OMC Name (e.g., Shell, Total, PSO)", "PSO")
    pmg = st.sidebar.number_input("PMG Sales (Kls)", min_value=0, value=100)
    hsd = st.sidebar.number_input("HSD Sales (Kls)", min_value=0, value=150)
    hobc = st.sidebar.number_input("HOBC Sales (Kls)", min_value=0, value=20)

    if st.sidebar.button("➕ Add Station"):
        st.session_state.elements.append(
            {
                "type": entry_type,
                "name": name,
                "omc": omc,
                "lat": lat,
                "lon": lon,
                "pmg": pmg,
                "hsd": hsd,
                "hobc": hobc,
            }
        )
        st.sidebar.success(f"Added {name}")
else:
    if st.sidebar.button("➕ Add Landmark"):
        st.session_state.elements.append(
            {
                "type": entry_type,
                "name": name,
                "omc": "N/A",
                "lat": lat,
                "lon": lon,
                "pmg": 0,
                "hsd": 0,
                "hobc": 0,
            }
        )
        st.sidebar.success(f"Added Landmark {name}")


# --- 2. MATH HELPERS ---
def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371.0  # Earth radius in km
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(dlon / 2) ** 2
    )
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# --- 3. MAP RENDERING ENGINE (FOLIUM) ---
proposed = [
    e for e in st.session_state.elements if e["type"] == "Proposed Station"
]
center_lat = proposed[0]["lat"] if proposed else 24.8607
center_lon = proposed[0]["lon"] if proposed else 67.0011

m = folium.Map(
    location=[center_lat, center_lon], zoom_start=12, tiles="CartoDB dark_matter"
)

folium.Circle(
    radius=10000,
    location=[center_lat, center_lon],
    color="#00D2FF",
    fill=True,
    fill_opacity=0.05,
    popup="10 km Trade Boundary",
).add_to(m)

folium.Circle(
    radius=5000,
    location=[center_lat, center_lon],
    color="#FFD166",
    fill=True,
    fill_opacity=0.08,
    popup="5 km Primary Zone",
).add_to(m)

for el in st.session_state.elements:
    if "Station" in el["type"]:
        is_p = el["type"] == "Proposed Station"
        border_color = "#00FF66" if is_p else "#00D2FF"
        bg_color = "#0F172A" if is_p else "#1E293B"

        html_card = f"""
        <div style="font-family: Arial; width: 170px; padding: 6px; border:2px solid {border_color}; background-color:{bg_color}; color:white; border-radius:6px; box-shadow: 2px 2px 6px rgba(0,0,0,0.5);">
            <div style="font-weight:bold; font-size:11px; text-align:center;">{el['name']}</div>
            <div style="background-color:#334155; text-align:center; font-size:9px; margin:3px 0; padding:2px; font-weight:bold; color:#FFD166;">★ {el['omc']} LOGO ★</div>
            <div style="font-size:9.5px; text-align:center; line-height:1.3;">
                PMG: <b>{el['pmg']} Kl</b> | HSD: <b>{el['hsd']} Kl</b><br>
                HOBC: <b>{el['hobc']} Kl</b>
            </div>
        </div>
        """
        folium.Marker(
            location=[el["lat"], el["lon"]], icon=folium.DivIcon(html=html_card)
        ).add_to(m)

        if not is_p and proposed:
            dist = calculate_distance(
                proposed[0]["lat"], proposed[0]["lon"], el["lat"], el["lon"]
            )
            folium.PolyLine(
                locations=[
                    [proposed[0]["lat"], proposed[0]["lon"]],
                    [el["lat"], el["lon"]],
                ],
                color="#00FF66",
                weight=2,
                opacity=0.8,
                tooltip=f"Distance: {dist:.2f} km",
            ).add_to(m)
    else:
        folium.Marker(
            location=[el["lat"], el["lon"]],
            popup=el["name"],
            icon=folium.Icon(color="red", icon="info-sign"),
        ).add_to(m)


# --- 4. NATIVE PYTHON MAP RENDERER FOR EXCEL (NO BROWSER NEEDED) ---
def generate_static_map_image(elements_data) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(10, 8), dpi=150)
    fig.patch.set_facecolor("#0F172A")
    ax.set_facecolor("#0F172A")

    proposed_list = [
        e for e in elements_data if e["type"] == "Proposed Station"
    ]
    p_lat = proposed_list[0]["lat"] if proposed_list else 24.8607
    p_lon = proposed_list[0]["lon"] if proposed_list else 67.0011

    # Draw 10 km and 5 km radius circles
    c10 = plt.Circle(
        (p_lon, p_lat),
        0.09,
        color="#00D2FF",
        fill=True,
        alpha=0.1,
        linestyle="--",
    )
    c5 = plt.Circle(
        (p_lon, p_lat),
        0.045,
        color="#FFD166",
        fill=True,
        alpha=0.15,
        linestyle=":",
    )
    ax.add_patch(c10)
    ax.add_patch(c5)

    for el in elements_data:
        e_lat, e_lon = el["lat"], el["lon"]
        if "Station" in el["type"]:
            is_p = el["type"] == "Proposed Station"
            color = "#00FF66" if is_p else "#00D2FF"
            ax.scatter(e_lon, e_lat, color=color, s=150, zorder=5)

            label = f"{el['name']}\n({el['omc']})\nPMG:{el['pmg']} HSD:{el['hsd']} HOBC:{el['hobc']}"
            ax.text(
                e_lon,
                e_lat + 0.005,
                label,
                color="white",
                fontsize=7,
                ha="center",
                bbox=dict(
                    boxstyle="round,pad=0.3",
                    facecolor="#1E293B",
                    edgecolor=color,
                ),
                zorder=6,
            )

            if not is_p and proposed_list:
                ax.plot(
                    [p_lon, e_lon],
                    [p_lat, e_lat],
                    color="#00FF66",
                    linestyle="--",
                    linewidth=1.5,
                )
                dist = calculate_distance(p_lat, p_lon, e_lat, e_lon)
                mid_lon, mid_lat = (p_lon + e_lon) / 2, (p_lat + e_lat) / 2
                ax.text(
                    mid_lon,
                    mid_lat,
                    f"{dist:.1f} km",
                    color="#00FF66",
                    fontsize=6,
                    fontweight="bold",
                    bbox=dict(boxstyle="square,pad=0.1", facecolor="#0F172A"),
                )
        else:
            ax.scatter(e_lon, e_lat, color="#FF4D4D", s=100, zorder=5)
            ax.text(
                e_lon,
                e_lat + 0.004,
                el["name"],
                color="#FF4D4D",
                fontsize=7,
                ha="center",
            )

    ax.set_title(
        "PETROL STATION TRADE AREA MAP (10 KM RADIUS)",
        color="white",
        fontsize=12,
        fontweight="bold",
    )
    ax.axis("off")
    plt.tight_layout()

    img_buf = io.BytesIO()
    plt.savefig(img_buf, format="png", dpi=150, facecolor=fig.get_facecolor())
    plt.close()
    img_buf.seek(0)
    return img_buf


def create_excel_report(elements_data) -> io.BytesIO:
    wb = Workbook()

    # Map Tab
    ws_map = wb.active
    ws_map.title = "Trade Area Map"
    ws_map.views.sheetView[0].showGridLines = True

    ws_map.merge_cells("A1:G2")
    title_cell = ws_map["A1"]
    title_cell.value = "PETROL STATION TRADE AREA REPORT (10 KM RADIUS)"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(
        start_color="1E293B", end_color="1E293B", fill_type="solid"
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    map_img_buf = generate_static_map_image(elements_data)
    img = OpenPyxlImage(map_img_buf)
    img.width = 750
    img.height = 600
    ws_map.add_image(img, "A4")

    # Data Tab
    ws_data = wb.create_sheet(title="Station Data")
    ws_data.views.sheetView[0].showGridLines = True

    headers = [
        "Type",
        "Station / Landmark Name",
        "OMC Brand",
        "PMG (Kls)",
        "HSD (Kls)",
        "HOBC (Kls)",
        "Total Sales (Kls)",
    ]
    ws_data.append(headers)

    header_fill = PatternFill(
        start_color="0F172A", end_color="0F172A", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, el in enumerate(elements_data, start=2):
        total_sales = (
            el.get("pmg", 0) + el.get("hsd", 0) + el.get("hobc", 0)
            if "Station" in el["type"]
            else 0
        )
        ws_data.append(
            [
                el["type"],
                el["name"],
                el.get("omc", "N/A"),
                el.get("pmg", 0) if "Station" in el["type"] else "-",
                el.get("hsd", 0) if "Station" in el["type"] else "-",
                el.get("hobc", 0) if "Station" in el["type"] else "-",
                total_sales if "Station" in el["type"] else "-",
            ]
        )

        for col_num in range(1, len(headers) + 1):
            c = ws_data.cell(row=row_idx, column=col_num)
            c.border = thin_border

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


# --- 5. MAIN DISPLAY & EXPORT INTERFACE ---
col1, col2 = st.columns([3, 1])

with col1:
    st_folium(m, width=900, height=650)

with col2:
    st.subheader("📊 Station Inventory")
    df = pd.DataFrame(st.session_state.elements)
    st.dataframe(
        df[["name", "type", "omc", "pmg", "hsd", "hobc"]], hide_index=True
    )

    st.markdown("---")
    st.subheader("📥 Report Generation")

    if st.button("Generate Excel Report"):
        with st.spinner("Building Excel sheet and rendering report map..."):
            excel_bytes = create_excel_report(st.session_state.elements)

            st.download_button(
                label="💾 Download Excel (.xlsx)",
                data=excel_bytes,
                file_name="Petrol_Station_Trade_Area_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.success("Report ready!")

    if st.button("Clear All Data"):
        st.session_state.elements = []
        st.rerun()
